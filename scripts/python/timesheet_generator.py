#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
工作日志生成器 - 按 Timesheet-202512.xlsx 模板 + 合并单元格 + Arial 14pt 居中
"""

import datetime
import os
import json
import argparse
import re
from pathlib import Path
from unicodedata import east_asian_width
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ========================================
# 🔧 配置区
# python3 timesheet_generator.py && ls -lh 2026-01/*.xlsx | tail -1
# ========================================

YEAR = 2026
MONTH = 1
PERSON_NAME = "彭聪"
PSP_COMPANY_VALUE = "英特利普（上海）信息技术有限公司"
REPORT_BASE_DIR = Path.home() / "Documents" / "work" / "月报"
OUTPUT_DIR = None
WORK_LOG_DIR = None
INCLUDE_ALL_LOGGED_DAYS = False
LOG_SOURCE_FIELDS = ["ailog"]
ARTIFACTS_SOURCE_TEMPLATE = "artifacts/timesheet-{year}{month:02d}-产物清单.txt"
DYNAMIC_DETAIL_WIDTH = True
DETAIL_WIDTH_MAX_MULTIPLIER = 4

COLUMN_WIDTHS = [22, 13, 20.8, 45, 15, 13]  # Week, Date, Project, Detail, Hours, Approval
ROW_HEIGHT = 34
DETAIL_ROW_BASE_HEIGHT = ROW_HEIGHT
DETAIL_ROW_LINE_HEIGHT = 18

PSP_NAME_LABEL = "PSP Name"
PSP_COMPANY_LABEL = "PSP Company"
OWNER_LABEL = "Starbucks Owner"
APPROVER = "Liangyu Chen"

PROJECT_NAME = "星巴克研发效能平台"
DEFAULT_HOURS = 8

CHINA_HOLIDAYS = set()  # 2025年12月无法定节假日
WEEKEND_DAYS = {5, 6}
# 当有周末补班（例：调休为工作日），放入 WORKING_EXCEPTIONS
WORKING_EXCEPTIONS = set()

# 是否优先使用线上接口获取中国节假日（若网络不可用会回退到硬编码列表）
USE_HOLIDAY_API = True

# 可编辑的 2026 年节假日（回退用，用户可按需修改）
# 请填写为 'YYYY-MM-DD' 字符串，函数会转换为 datetime.date
HARDCODED_HOLIDAYS_2026 = [
    # 元旦：1月1-3日
    '2026-01-01', '2026-01-02', '2026-01-03',
    # 春节：2月15-23日
    '2026-02-15', '2026-02-16', '2026-02-17', '2026-02-18', '2026-02-19',
    '2026-02-20', '2026-02-21', '2026-02-22', '2026-02-23',
    # 清明节：4月4-6日
    '2026-04-04', '2026-04-05', '2026-04-06',
    # 劳动节：5月1-5日
    '2026-05-01', '2026-05-02', '2026-05-03', '2026-05-04', '2026-05-05',
    # 端午节：6月19-21日
    '2026-06-19', '2026-06-20', '2026-06-21',
    # 中秋节：9月25-27日
    '2026-09-25', '2026-09-26', '2026-09-27',
    # 国庆节：10月1-7日
    '2026-10-01', '2026-10-02', '2026-10-03', '2026-10-04',
    '2026-10-05', '2026-10-06', '2026-10-07',
]
# 可编辑的 2026 年补班（周末变为工作日），格式同上
HARDCODED_WORKING_2026 = [
    # 元旦调休：1月4日（周日）上班
    '2026-01-04',
    # 春节调休：2月14日（周六）、2月28日（周六）上班
    '2026-02-14', '2026-02-28',
    # 劳动节调休：5月9日（周六）上班
    '2026-05-09',
    # 中秋节/国庆节调休：9月20日（周日）、10月10日（周六）上班
    '2026-09-20', '2026-10-10',
]

# ========================================
# 📅 工具函数
# ========================================

def parse_dotenv_file(env_path):
    if not env_path.exists():
        return

    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = value


def load_dotenv(env_file=None):
    repo_root = Path(__file__).resolve().parents[1]
    env_paths = [Path(env_file).expanduser()] if env_file else [
        repo_root / ".agents" / "skills" / "monthly-json-timesheet" / ".env",
        repo_root / ".env",
        Path.cwd() / ".env",
    ]

    seen = set()
    for env_path in env_paths:
        env_path = env_path.resolve()
        if env_path in seen:
            continue
        seen.add(env_path)
        parse_dotenv_file(env_path)


def env_bool(name, default):
    value = os.environ.get(name)
    if value is None:
        return default
    return value.strip().lower() in {"1", "true", "yes", "y", "on"}


def env_int(name, default):
    value = os.environ.get(name)
    if value is None or not value.strip():
        return default
    try:
        return int(value)
    except ValueError:
        return default


def apply_env_config():
    global PERSON_NAME, PSP_COMPANY_VALUE, REPORT_BASE_DIR, OUTPUT_DIR, LOG_SOURCE_FIELDS
    global APPROVER, PROJECT_NAME, DEFAULT_HOURS, USE_HOLIDAY_API
    global ARTIFACTS_SOURCE_TEMPLATE, DYNAMIC_DETAIL_WIDTH, DETAIL_WIDTH_MAX_MULTIPLIER

    PERSON_NAME = os.environ.get("TIMESHEET_PSP_NAME", PERSON_NAME)
    PSP_COMPANY_VALUE = os.environ.get("TIMESHEET_PSP_COMPANY", PSP_COMPANY_VALUE)
    APPROVER = os.environ.get("TIMESHEET_APPROVER", APPROVER)
    PROJECT_NAME = os.environ.get("TIMESHEET_PROJECT_NAME", PROJECT_NAME)
    REPORT_BASE_DIR = Path(os.environ.get("TIMESHEET_OUTPUT_BASE", str(REPORT_BASE_DIR))).expanduser()
    OUTPUT_DIR = os.environ.get("TIMESHEET_OUTPUT_DIR") or OUTPUT_DIR
    LOG_SOURCE_FIELDS = normalize_source_fields(os.environ.get("TIMESHEET_SOURCE_FIELDS", ",".join(LOG_SOURCE_FIELDS)))
    ARTIFACTS_SOURCE_TEMPLATE = os.environ.get("TIMESHEET_ARTIFACTS_SOURCE_TEMPLATE", ARTIFACTS_SOURCE_TEMPLATE)
    DEFAULT_HOURS = env_int("TIMESHEET_DEFAULT_HOURS", DEFAULT_HOURS)
    USE_HOLIDAY_API = env_bool("TIMESHEET_USE_HOLIDAY_API", USE_HOLIDAY_API)
    DYNAMIC_DETAIL_WIDTH = env_bool("TIMESHEET_DYNAMIC_DETAIL_WIDTH", DYNAMIC_DETAIL_WIDTH)
    DETAIL_WIDTH_MAX_MULTIPLIER = env_int("TIMESHEET_DETAIL_WIDTH_MAX_MULTIPLIER", DETAIL_WIDTH_MAX_MULTIPLIER)


def get_work_log_root():
    if WORK_LOG_DIR:
        return Path(WORK_LOG_DIR).expanduser()
    return Path(__file__).parent


def get_work_log_month_dir(year, month):
    return get_work_log_root() / f"{year}-{month:02d}"


def get_output_dir(year, month):
    if OUTPUT_DIR:
        output_dir = Path(OUTPUT_DIR).expanduser()
    else:
        output_dir = get_work_log_month_dir(year, month)
    output_dir.mkdir(parents=True, exist_ok=True)
    return output_dir


def normalize_source_fields(source_fields):
    if isinstance(source_fields, str):
        source_fields = [field.strip() for field in source_fields.split(",")]
    return [field for field in source_fields if field]


def display_width(text):
    width = 0
    for char in str(text):
        width += 2 if east_asian_width(char) in {"F", "W"} else 1
    return width


def calculate_detail_column_width(monthly_data):
    base_width = COLUMN_WIDTHS[3]
    if not DYNAMIC_DETAIL_WIDTH:
        return base_width

    max_line_width = base_width
    for detail in monthly_data.values():
        for line in str(detail).splitlines():
            max_line_width = max(max_line_width, display_width(line) + 2)

    max_allowed = base_width * max(1, DETAIL_WIDTH_MAX_MULTIPLIER)
    return min(max_line_width, max_allowed)


def calculate_detail_row_height(detail):
    item_count = len([line for line in str(detail or "").splitlines() if line.strip()])
    item_count = max(1, item_count)
    return max(DETAIL_ROW_BASE_HEIGHT, item_count * DETAIL_ROW_LINE_HEIGHT)


def collect_daily_data(year, month, source_fields):
    month_dir = get_work_log_month_dir(year, month)
    daily_data = {}
    source_fields = normalize_source_fields(source_fields)

    for daily_file in sorted(month_dir.glob(f"{year}-{month:02d}-*.json")):
        if not re.match(r'^\d{4}-\d{2}-\d{2}\.json$', daily_file.name):
            continue
        try:
            with open(daily_file, 'r', encoding='utf-8') as f:
                daily_log = json.load(f)

            date = daily_log.get('date', '')
            tasks = []
            for field in source_fields:
                value = daily_log.get(field)
                if isinstance(value, list) and value:
                    tasks = [str(item).strip() for item in value if str(item).strip()]
                    if tasks:
                        break

            if date and tasks:
                daily_data[date.replace('-', '')] = '\n'.join(tasks)
        except Exception as e:
            print(f"  ⚠️  跳过文件 {daily_file.name}: {e}")

    return daily_data


def infer_daily_hours(daily_log):
    notes = str(daily_log.get("notes", ""))
    if "请假半天" in notes or "半天请假" in notes:
        return DEFAULT_HOURS / 2
    return DEFAULT_HOURS


def collect_daily_hours(year, month):
    month_dir = get_work_log_month_dir(year, month)
    daily_hours = {}

    for daily_file in sorted(month_dir.glob(f"{year}-{month:02d}-*.json")):
        if not re.match(r'^\d{4}-\d{2}-\d{2}\.json$', daily_file.name):
            continue
        try:
            daily_log = json.loads(daily_file.read_text(encoding="utf-8"))
        except Exception:
            continue
        date = daily_log.get("date", "")
        if date:
            daily_hours[date.replace("-", "")] = infer_daily_hours(daily_log)

    return daily_hours


def load_monthly_data(year, month, source_fields=None):
    """
    加载月度汇总 JSON 数据
    
    返回字典：{"20260101": "任务1 & 任务2", ...}
    """
    source_fields = normalize_source_fields(source_fields or LOG_SOURCE_FIELDS)
    daily_data = collect_daily_data(year, month, source_fields)
    if daily_data:
        print(f"✅ 已从每日 JSON 加载: {len(daily_data)} 个工作日记录，字段优先级: {','.join(source_fields)}")
        return daily_data
    
    data_file = get_work_log_month_dir(year, month) / f"{year}-{month:02d}.json"
    if not data_file.exists():
        print(f"⚠️  未找到月度数据文件: {data_file}")
        return {}
    
    try:
        with open(data_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
        print(f"✅ 已加载月度数据: {len(data)} 个工作日记录")
        return data
    except Exception as e:
        print(f"❌ 读取月度数据失败: {e}")
        return {}

def is_workday(date):
    # 如果明确是补班日期，则为工作日
    if date in WORKING_EXCEPTIONS:
        return True
    # 周末且非补班则非工作日
    if date.weekday() in WEEKEND_DAYS:
        return False
    # 法定节假日
    if date in CHINA_HOLIDAYS:
        return False
    return True


def load_china_holidays(year, use_api=USE_HOLIDAY_API):
    """尝试从线上接口加载节假日与补班信息；失败时使用硬编码回退。

    返回 (holidays_set, working_exceptions_set)
    """
    import datetime as _dt
    holidays = set()
    working = set()

    if use_api:
        try:
            import urllib.request, json
            url = f"https://timor.tech/api/holiday/year/{year}"
            print(f"🌐 正在从接口获取 {year} 年节假日数据...")
            with urllib.request.urlopen(url, timeout=10) as resp:
                data = json.load(resp)

            # 解析 timor.tech API 返回的数据结构
            # 格式：{"code": 0, "holiday": {"MM-DD": {"holiday": true/false, "date": "YYYY-MM-DD", ...}}}
            if data.get('code') == 0 and 'holiday' in data:
                for day_key, day_info in data['holiday'].items():
                    if isinstance(day_info, dict) and 'date' in day_info:
                        try:
                            dt = _dt.date.fromisoformat(day_info['date'])
                            # holiday: true 表示放假，false 表示补班
                            if day_info.get('holiday') is True:
                                holidays.add(dt)
                            elif day_info.get('holiday') is False:
                                working.add(dt)
                        except Exception:
                            continue

            # 检查是否成功解析到数据
            if holidays or working:
                print(f"✅ 成功从接口获取：{len(holidays)} 个假日，{len(working)} 个补班日")
            else:
                raise ValueError('接口未返回有效数据，使用硬编码回退')

        except Exception as e:
            print(f"⚠️ 接口调用失败 ({e})，使用硬编码数据")
            # 回退到硬编码（只处理 2026 年）
            if year == 2026:
                for s in HARDCODED_HOLIDAYS_2026:
                    try:
                        holidays.add(_dt.date.fromisoformat(s))
                    except Exception:
                        pass
                for s in HARDCODED_WORKING_2026:
                    try:
                        working.add(_dt.date.fromisoformat(s))
                    except Exception:
                        pass
                print(f"✅ 使用硬编码数据：{len(holidays)} 个假日，{len(working)} 个补班日")
    else:
        # 直接使用硬编码
        if year == 2026:
            for s in HARDCODED_HOLIDAYS_2026:
                try:
                    holidays.add(_dt.date.fromisoformat(s))
                except Exception:
                    pass
            for s in HARDCODED_WORKING_2026:
                try:
                    working.add(_dt.date.fromisoformat(s))
                except Exception:
                    pass
            print(f"✅ 使用硬编码数据：{len(holidays)} 个假日，{len(working)} 个补班日")

    return holidays, working


def parse_daily_gitlog_markdown(md_file):
    """解析 collect-monthly-timesheet 生成的工作日报清单.md。"""
    if not md_file.exists():
        return {}

    daily = {}
    current_date = None
    for raw_line in md_file.read_text(encoding='utf-8').splitlines():
        line = raw_line.strip()
        date_match = re.match(r'^(\d{4})/(\d{2})/(\d{2})$', line)
        if date_match:
            current_date = f"{date_match.group(1)}-{date_match.group(2)}-{date_match.group(3)}"
            daily.setdefault(current_date, [])
            continue

        if current_date and line.startswith('- '):
            item = line[2:].strip()
            if item:
                daily[current_date].append(item)

    return daily


def merge_gitlog_into_daily_json(year, month):
    """把 gitlog/工作日报清单.md 合并进每日 JSON，并补齐新字段。"""
    month_dir = get_work_log_month_dir(year, month)
    gitlog_file = month_dir / "gitlog" / "工作日报清单.md"
    daily_gitlog = parse_daily_gitlog_markdown(gitlog_file)

    if not daily_gitlog:
        return

    month_dir.mkdir(exist_ok=True)
    updated = 0
    existing_dates = {
        p.stem
        for p in month_dir.glob(f"{year}-{month:02d}-*.json")
        if re.match(r'^\d{4}-\d{2}-\d{2}$', p.stem)
    }
    all_dates = sorted(existing_dates | set(daily_gitlog.keys()))

    for date_str in all_dates:
        items = daily_gitlog.get(date_str, [])
        json_file = month_dir / f"{date_str}.json"
        if json_file.exists():
            try:
                daily_log = json.loads(json_file.read_text(encoding='utf-8'))
            except Exception:
                daily_log = {}
        else:
            daily_log = {}

        daily_log.setdefault("date", date_str)
        daily_log.setdefault("completed", [])
        daily_log.setdefault("plan", [])
        daily_log.setdefault("blockers", [])
        daily_log.setdefault("notes", "")
        daily_log["gitlog"] = items if items else daily_log.get("gitlog", [])
        daily_log.setdefault("ailog", [])
        daily_log.setdefault("origin_url", [])

        json_file.write_text(json.dumps(daily_log, ensure_ascii=False, indent=2) + "\n", encoding='utf-8')
        updated += 1

    print(f"✅ 已合并 GitLog 到每日 JSON: {updated} 天")


def normalize_repo_url(origin_url):
    origin_url = origin_url.strip()
    if origin_url.startswith("git@scm.starbucks.com:"):
        origin_url = origin_url.replace("git@scm.starbucks.com:", "https://scm.starbucks.com/", 1)
        if origin_url.endswith(".git"):
            origin_url = origin_url[:-4]
    return origin_url


def dedupe_preserve_order(items):
    seen = set()
    deduped = []
    for item in items:
        if item in seen:
            continue
        seen.add(item)
        deduped.append(item)
    return deduped


def resolve_artifacts_source_file(year, month):
    source = ARTIFACTS_SOURCE_TEMPLATE.format(year=year, month=month)
    source_path = Path(source).expanduser()
    if not source_path.is_absolute():
        source_path = Path(__file__).resolve().parents[1] / source_path
    return source_path


def read_artifact_urls_from_text(source_file):
    urls = []
    for raw_line in source_file.read_text(encoding='utf-8').splitlines():
        line = raw_line.strip()
        if not line.startswith("- "):
            continue
        if "|" in line:
            origin_url = line.split("|", 1)[1].strip()
        else:
            origin_url = line[2:].strip()
        origin_url = normalize_repo_url(origin_url)
        if origin_url:
            urls.append(origin_url)
    return dedupe_preserve_order(urls)


def read_artifact_urls_from_tsv(source_file):
    urls = []
    for line in source_file.read_text(encoding='utf-8').splitlines():
        if not line.strip():
            continue
        cols = line.split('\t')
        if cols[0] == "repo_path":
            continue
        origin_url = normalize_repo_url((cols + [""] * 3)[2])
        if origin_url:
            urls.append(origin_url)
    return dedupe_preserve_order(urls)


def generate_artifacts_excel(year, month):
    """生成交付物 Excel。优先从 gitlog/产物清单.tsv 读取。"""
    month_dir = get_work_log_month_dir(year, month)
    fallback_file = month_dir / "gitlog" / "产物清单.tsv"
    artifacts_file = fallback_file if fallback_file.exists() else resolve_artifacts_source_file(year, month)

    if artifacts_file.exists():
        if artifacts_file.suffix == ".tsv":
            artifact_urls = read_artifact_urls_from_tsv(artifacts_file)
        else:
            artifact_urls = read_artifact_urls_from_text(artifacts_file)
        print(f"✅ 已加载交付物来源: {artifacts_file}")
    elif fallback_file.exists():
        artifact_urls = read_artifact_urls_from_tsv(fallback_file)
        print(f"⚠️  未找到交付物来源: {artifacts_file}，改用 {fallback_file}")
    else:
        print(f"⚠️  未找到产物清单: {artifacts_file}")
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    body_font = Font(name='宋体', size=11)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws.cell(row=1, column=1, value="代码仓库地址：")
    ws.cell(row=1, column=1).font = body_font
    ws.cell(row=1, column=1).alignment = left

    row_index = 2
    for origin_url in artifact_urls:
        cell = ws.cell(row=row_index, column=1, value=origin_url)
        cell.font = body_font
        cell.alignment = left
        row_index += 1

    ws.column_dimensions["A"].width = 61.83203125
    for row in range(1, row_index):
        ws.row_dimensions[row].height = 28

    output_file = get_output_dir(year, month) / f"交付物_{PERSON_NAME}_{year}{month:02d}.xlsx"
    wb.save(output_file)
    print(f"✅ 已生成交付物：{output_file}")
    return output_file

# ========================================
# 📊 主函数
# ========================================

def create_psp_work_log_excel():
    merge_gitlog_into_daily_json(YEAR, MONTH)
    generate_artifacts_excel(YEAR, MONTH)

    # 加载中国节假日与补班数据（优先尝试接口，失败回退到硬编码）
    global CHINA_HOLIDAYS, WORKING_EXCEPTIONS
    CHINA_HOLIDAYS, WORKING_EXCEPTIONS = load_china_holidays(YEAR, use_api=USE_HOLIDAY_API)
    
    # 加载月度工作日志数据
    monthly_data = load_monthly_data(YEAR, MONTH, LOG_SOURCE_FIELDS)
    daily_hours = collect_daily_hours(YEAR, MONTH)

    wb = Workbook()
    ws = wb.active
    ws.title = "工作日志"
    
    # 统一字体和对齐
    font_12 = Font(name='宋体', size=12)
    bold_font_12 = Font(name='宋体', size=12, bold=True)
    font_14 = Font(name='Arial', size=14, )
    bold_font_14 = Font(name='Arial', size=14, bold=True)
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    detail_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    header_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 写入顶部三行（准备合并）
    ws.cell(row=1, column=1, value=PSP_NAME_LABEL).font = bold_font_14
    ws.cell(row=2, column=1, value=PSP_COMPANY_LABEL).font = bold_font_14
    ws.cell(row=3, column=1, value=OWNER_LABEL).font = bold_font_14
    ws.cell(row=1, column=3, value=PERSON_NAME).font = bold_font_14
    ws.cell(row=2, column=3, value=PSP_COMPANY_VALUE).font = bold_font_14
    ws.cell(row=3, column=3, value=APPROVER).font = bold_font_14

    # 设置顶部三行样式
    for row in [1, 2, 3]:
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            # cell.border = border
            cell.alignment = center_alignment
            # 保留我们放置值的位置（左侧三行首列与右上角 F1），其余清空
            if not (
                (row == 1 and col == 1)
                or (row == 2 and col == 1)
                or (row == 3 and col == 1)
                or (row == 1 and col == 3)
                or (row == 2 and col == 3)
                or (row == 3 and col == 3)
            ):
                cell.value = ""
        ws.row_dimensions[row].height = ROW_HEIGHT

    # 合并顶部单元格：左侧 A1:E1、A2:E2、A3:E3，右侧竖向 F1:F3（签名/审批）
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=2)
    ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=4)
    ws.merge_cells(start_row=2, start_column=3, end_row=2, end_column=4)
    ws.merge_cells(start_row=3, start_column=3, end_row=3, end_column=4)

    # 合并后确保左上单元格样式保持居中（需要对合并的左上角单元格设置样式）
    ws.cell(row=1, column=1).alignment = center_alignment
    ws.cell(row=2, column=1).alignment = center_alignment
    ws.cell(row=3, column=1).alignment = center_alignment
    ws.cell(row=1, column=3).alignment = center_alignment
    ws.cell(row=2, column=3).alignment = center_alignment
    ws.cell(row=3, column=3).alignment = center_alignment
    
    # 表头（第5行）
    headers = ['Week', 'Date', 'Project Name', 'Detail Description', 'Working Hours', 'Approval']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = bold_font_12
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    ws.row_dimensions[5].height = ROW_HEIGHT
    
    # 生成当月工作日
    start_date = datetime.date(YEAR, MONTH, 1)
    end_date = (start_date.replace(day=28) + datetime.timedelta(days=4)).replace(day=1) - datetime.timedelta(days=1)
    
    work_days = []
    current_date = start_date
    while current_date <= end_date:
        if is_workday(current_date):
            work_days.append(current_date)
        current_date += datetime.timedelta(days=1)

    if INCLUDE_ALL_LOGGED_DAYS:
        existing = {day for day in work_days}
        for date_key, detail in monthly_data.items():
            if not str(detail).strip():
                continue
            try:
                extra = datetime.date(
                    int(date_key[0:4]),
                    int(date_key[4:6]),
                    int(date_key[6:8]),
                )
            except (ValueError, IndexError):
                continue
            if extra not in existing:
                work_days.append(extra)
                existing.add(extra)
        work_days.sort()
    
    # 按连续工作日分组（连着上班的算一周）
    weeks_data = []
    if work_days:
        current_week = [work_days[0]]
        for i in range(1, len(work_days)):
            # 检查当前日期是否与前一天连续（相差1天）
            if (work_days[i] - work_days[i-1]).days == 1:
                current_week.append(work_days[i])
            else:
                # 不连续，开启新的一周
                weeks_data.append(current_week)
                current_week = [work_days[i]]
        # 添加最后一周
        weeks_data.append(current_week)

    # 设置列宽；D 列按明细最长行动态扩展，避免列表项被自动换行。
    column_widths = list(COLUMN_WIDTHS)
    column_widths[3] = calculate_detail_column_width(monthly_data)
    detail_column_width = column_widths[3]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # 写入数据行（从第6行开始）
    row = 6
    week_num = 1
    approval_start_row = row  # 记录 Approval 列起始行
    
    week_ranges = []  # 记录每个 Week 的起止行 [start, end]
    
    for days_in_week in weeks_data:
        week_label = f"Week{week_num}"
        week_start_row = row
        
        for day in days_in_week:
            date_key = day.strftime('%Y%m%d')
            detail_description = monthly_data.get(date_key, "")  # 从月度数据获取任务描述
            working_hours = daily_hours.get(date_key, DEFAULT_HOURS)
            
            ws.cell(row=row, column=1, value=week_label)
            ws.cell(row=row, column=2, value=date_key)
            ws.cell(row=row, column=3, value=PROJECT_NAME)
            ws.cell(row=row, column=4, value=detail_description)
            ws.cell(row=row, column=5, value=working_hours)
            ws.cell(row=row, column=6, value=APPROVER)
            
            # 应用样式
            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                cell.font = font_12
                cell.alignment = detail_alignment if col == 4 else center_alignment
                cell.border = border
            
            ws.row_dimensions[row].height = calculate_detail_row_height(detail_description)
            row += 1
        
        # 记录 Week 合并范围
        week_ranges.append((week_start_row, row - 1))
        week_num += 1
    
    approval_end_row = row - 1  # Approval 列结束行
    
    # 合计行
    total_hours = sum(daily_hours.get(day.strftime('%Y%m%d'), DEFAULT_HOURS) for day in work_days)
    total_days = total_hours / DEFAULT_HOURS
    if total_days == int(total_days):
        total_days = int(total_days)
    
    total_hours_row = row
    total_days_row = row + 1
    
    # Total Hours 行
    for col in range(1, 5):
        ws.cell(row=total_hours_row, column=col, value="Total Hours")
    ws.cell(row=total_hours_row, column=5, value=total_hours)
    ws.cell(row=total_hours_row, column=6, value="Hours")
    
    # Total Days 行
    for col in range(1, 5):
        ws.cell(row=total_days_row, column=col, value="Total Days")
    ws.cell(row=total_days_row, column=5, value=total_days)
    ws.cell(row=total_days_row, column=6, value="Days")
    
    # 合计行样式
    for r in [total_hours_row, total_days_row]:
        for col in range(1, 7):
            cell = ws.cell(row=r, column=col)
            cell.font = bold_font_12
            cell.alignment = center_alignment
            cell.border = border
            if col <= 6:
                cell.fill = header_fill
        ws.row_dimensions[r].height = ROW_HEIGHT
    
    # ================================
    # 🔧 执行单元格合并
    # ================================
    
    # 1. 合并 Week 列
    for start_row, end_row in week_ranges:
        if start_row != end_row:  # 多于一行才合并
            ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
    
    # 2. 合并 Approval 列（整列合并）
    if approval_start_row <= approval_end_row:
        ws.merge_cells(start_row=approval_start_row, start_column=6, end_row=approval_end_row, end_column=6)
    
    # 3. 合并 Total Hours 的 A:D
    ws.merge_cells(start_row=total_hours_row, start_column=1, end_row=total_hours_row, end_column=4)
    
    # 4. 合并 Total Days 的 A:D
    ws.merge_cells(start_row=total_days_row, start_column=1, end_row=total_days_row, end_column=4)
    
    # 重新设置合并后单元格的样式（openpyxl 合并后需手动设样式）
    merged_cells = [
        (total_hours_row, 1),
        (total_days_row, 1),
    ]
    for r, c in merged_cells:
        cell = ws.cell(row=r, column=c)
        cell.font = bold_font_12
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    # 保存文件（按月份创建文件夹，避免覆盖）
    # 创建输出目录：outputs/YYYY-MM/
    output_dir = get_output_dir(YEAR, MONTH)
    
    filename = f"Timesheet-{PERSON_NAME}_{YEAR}{MONTH:02d}.xlsx"
    filepath = output_dir / filename
    
    wb.save(filepath)
    print(f"✅ 已生成：{filepath}")

# ========================================
# 🚀 运行
# ========================================

if __name__ == "__main__":
    load_dotenv()
    apply_env_config()

    parser = argparse.ArgumentParser(description="生成 Timesheet 和交付物清单")
    parser.add_argument("--year", type=int, default=YEAR)
    parser.add_argument("--month", type=int, default=MONTH)
    parser.add_argument("--output-base", default=str(REPORT_BASE_DIR), help="月报输出根目录，默认 ~/Documents/work/月报")
    parser.add_argument("--output-dir", default=None, help="指定完整输出目录；优先级高于 --output-base")
    parser.add_argument("--source-fields", default=",".join(LOG_SOURCE_FIELDS), help="日报字段优先级，逗号分隔，默认 ailog")
    parser.add_argument("--psp-name", default=PERSON_NAME)
    parser.add_argument("--psp-company", default=PSP_COMPANY_VALUE)
    parser.add_argument("--approver", default=APPROVER)
    parser.add_argument("--project-name", default=PROJECT_NAME)
    parser.add_argument("--artifacts-source-template", default=ARTIFACTS_SOURCE_TEMPLATE)
    parser.add_argument("--dynamic-detail-width", action=argparse.BooleanOptionalAction, default=DYNAMIC_DETAIL_WIDTH)
    parser.add_argument("--detail-width-max-multiplier", type=int, default=DETAIL_WIDTH_MAX_MULTIPLIER)
    parser.add_argument("--work-log-dir", default=None, help="日报 JSON 根目录，例如 ~/.work-logs")
    parser.add_argument(
        "--include-all-logged-days",
        action="store_true",
        help="包含所有有日志记录的非标准工作日（全日期工时表）",
    )
    args = parser.parse_args()
    YEAR = args.year
    MONTH = args.month
    REPORT_BASE_DIR = Path(args.output_base).expanduser()
    OUTPUT_DIR = args.output_dir
    WORK_LOG_DIR = args.work_log_dir
    INCLUDE_ALL_LOGGED_DAYS = args.include_all_logged_days
    LOG_SOURCE_FIELDS = normalize_source_fields(args.source_fields)
    PERSON_NAME = args.psp_name
    PSP_COMPANY_VALUE = args.psp_company
    APPROVER = args.approver
    PROJECT_NAME = args.project_name
    ARTIFACTS_SOURCE_TEMPLATE = args.artifacts_source_template
    DYNAMIC_DETAIL_WIDTH = args.dynamic_detail_width
    DETAIL_WIDTH_MAX_MULTIPLIER = args.detail_width_max_multiplier
    create_psp_work_log_excel()
