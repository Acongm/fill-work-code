#!/usr/bin/env python3
import argparse
import csv
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


def read_remarks(path: Optional[Path]) -> dict[str, str]:
    if path is None or not path.exists():
        return {}

    remarks: dict[str, str] = {}
    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.reader(handle, delimiter="\t")
        for row in reader:
            if len(row) < 2:
                continue
            key = row[0].strip()
            value = row[1].strip()
            if key:
                remarks[key] = value
    return remarks


def remark_for(repo_name: str, origin_url: str, remarks: dict[str, str]) -> str:
    candidates = [
        repo_name,
        origin_url,
        origin_url.rstrip("/").removesuffix(".git").split("/")[-1],
        origin_url.rstrip("/").removesuffix(".git").split(":")[-1].split("/")[-1],
    ]
    for key in candidates:
        if key in remarks:
            return remarks[key]
    return ""


def load_repo_rows(input_tsv: Path, remarks: dict[str, str]) -> list[tuple[str, str]]:
    rows: list[tuple[str, str]] = []
    seen: set[str] = set()
    with input_tsv.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.reader(handle, delimiter="\t")
        for row in reader:
            if len(row) < 3:
                continue
            repo_name = row[1].strip()
            origin_url = row[2].strip()
            if not origin_url or origin_url in seen:
                continue
            seen.add(origin_url)
            rows.append((origin_url, remark_for(repo_name, origin_url, remarks)))
    return rows


def render_artifact_list(rows: list[tuple[str, str]], output_xlsx: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    has_remarks = any(remark for _, remark in rows)
    ws.cell(1, 1, "代码仓库地址：")
    if has_remarks:
        ws.cell(1, 2, "备注")

    header_font = Font(name="SimSun", bold=True, size=12)
    normal_font = Font(name="SimSun", size=12)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")

    for index, (origin_url, remark) in enumerate(rows, start=2):
        ws.cell(index, 1, origin_url)
        ws.cell(index, 1).font = normal_font
        ws.cell(index, 1).alignment = Alignment(vertical="center")
        if has_remarks:
            ws.cell(index, 2, remark)
            ws.cell(index, 2).font = normal_font
            ws.cell(index, 2).alignment = Alignment(vertical="center", wrap_text=True)

    ws.column_dimensions[get_column_letter(1)].width = 70
    if has_remarks:
        ws.column_dimensions[get_column_letter(2)].width = 28

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Export monthly repository artifact list to xlsx")
    parser.add_argument("--input-tsv", required=True, help="TSV emitted by list-monthly-repos.sh")
    parser.add_argument("--output-xlsx", required=True, help="artifact list workbook path")
    parser.add_argument("--remarks-file", default="", help="optional TSV: repo name or URL, then remark")
    return parser


def main():
    args = build_parser().parse_args()
    remarks_file = Path(args.remarks_file) if args.remarks_file else None
    remarks = read_remarks(remarks_file)
    rows = load_repo_rows(Path(args.input_tsv), remarks)
    render_artifact_list(rows, Path(args.output_xlsx))
    print(f"Artifact list written: {args.output_xlsx} ({len(rows)} repositories)")


if __name__ == "__main__":
    main()
